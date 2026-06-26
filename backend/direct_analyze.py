import os
import json
import time
import openpyxl
from google import genai
from google.genai import types

# ✅ Paste your fresh API key here directly
GEMINI_API_KEY = "AQ.Ab8RN6JESwlWiij2hgjda_wpdLwXWAJy6WJJoR_KKOyiAfHg6w"

JPG_FOLDER = "bills_jpg"
OUTPUT_FILE = "../ml-models/bills_dataset.xlsx"

client = genai.Client(api_key=GEMINI_API_KEY)

skip_list = ["54361222224.jpg", "80421700005.jpg"]

# Load existing Excel or create new
if os.path.exists(OUTPUT_FILE):
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    already_done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            already_done.add(row[0])
    print(f"📋 Found existing dataset — {len(already_done)} bills already processed, skipping them.\n")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills Data"
    ws.append(["File", "Customer Name", "Consumer Number", "DISCOM",
               "City", "Monthly Units", "Bill Amount", "Per Unit Rate",
               "Billing Period", "Recommended KW", "Monthly Savings Rs",
               "System Cost Rs", "Payback Years", "25 Year Savings Rs"])
    already_done = set()

jpg_files = [f for f in os.listdir(JPG_FOLDER) if f.lower().endswith(".jpg")]
remaining = [f for f in jpg_files if f not in already_done and f not in skip_list]
print(f"Found {len(jpg_files)} total bills — {len(remaining)} remaining...\n")

success = 0
failed = 0

prompt = """
You are an expert at reading Indian electricity bills.
Extract the following from this bill image and return ONLY valid JSON:
{
    "customer_name": "<name>",
    "consumer_number": "<number>",
    "discom": "<company>",
    "monthly_units": <number>,
    "bill_amount": <number>,
    "per_unit_rate": <number>,
    "billing_period": "<month year>",
    "recommended_kw": <monthly_units/135 rounded to 0.5>,
    "monthly_generation_units": <recommended_kw*4.5*30>,
    "monthly_savings_rs": <monthly_generation*per_unit_rate>,
    "system_cost_rs": <recommended_kw*50000>,
    "payback_years": <system_cost/(monthly_savings*12)>,
    "savings_25_years_rs": <(monthly_savings*12*25)-system_cost>
}
"""

for filename in remaining:
    filepath = os.path.join(JPG_FOLDER, filename)
    try:
        with open(filepath, "rb") as f:
            image_data = f.read()

        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=[
                types.Content(role="user", parts=[
                    types.Part.from_bytes(data=image_data, mime_type="image/jpeg"),
                    types.Part.from_text(text=prompt)
                ])
            ]
        )

        text = response.text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]

        d = json.loads(text.strip())

        ws.append([
            filename,
            d.get("customer_name", ""),
            d.get("consumer_number", ""),
            d.get("discom", ""),
            "",  # City — fill manually
            d.get("monthly_units", ""),
            d.get("bill_amount", ""),
            d.get("per_unit_rate", ""),
            d.get("billing_period", ""),
            d.get("recommended_kw", ""),
            d.get("monthly_savings_rs", ""),
            d.get("system_cost_rs", ""),
            d.get("payback_years", ""),
            d.get("savings_25_years_rs", "")
        ])
        wb.save(OUTPUT_FILE)
        success += 1
        print(f"✅ {filename} — ₹{d.get('bill_amount')} | {d.get('monthly_units')} units")
        time.sleep(4)  # delay between requests

    except Exception as e:
        failed += 1
        print(f"❌ {filename} — {str(e)}")

print(f"\n✅ Done! {success} saved, {failed} failed.")
print(f"📊 Saved to: {OUTPUT_FILE}")