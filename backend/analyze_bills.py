import requests
import os
import openpyxl
import time

JPG_FOLDER = "bills_jpg"
OUTPUT_FILE = "../ml-models/bills_dataset.xlsx"
API_URL = "http://localhost:8000/api/analyze-bill"

# Load existing Excel or create new one
if os.path.exists(OUTPUT_FILE):
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    # Get already processed filenames from column 1
    already_done = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            already_done.add(row[0])
    print(f"📋 Found existing dataset — {len(already_done)} bills already processed, skipping them.\n")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills Data"
    headers = [
        "File", "Customer Name", "Consumer Number", "DISCOM",
        "City", "Monthly Units", "Bill Amount", "Per Unit Rate",
        "Billing Period", "Recommended KW", "Monthly Savings Rs",
        "System Cost Rs", "Payback Years", "25 Year Savings Rs"
    ]
    ws.append(headers)
    already_done = set()

success = 0
failed = 0
MAX_RETRIES = 3
RETRY_DELAY = 60  # seconds between retries

jpg_files = [f for f in os.listdir(JPG_FOLDER) if f.lower().endswith(".jpg")]
skip_list = ["54361222224.jpg", "80421700005.jpg"]
remaining = [f for f in jpg_files if f not in already_done and f not in skip_list]
print(f"Found {len(jpg_files)} total bills — {len(remaining)} remaining to process...\n")

for filename in remaining:
    filepath = os.path.join(JPG_FOLDER, filename)
    attempt = 0
    processed = False

    while attempt < MAX_RETRIES and not processed:
        try:
            if attempt > 0:
                print(f"   ⏳ Retrying {filename} (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(RETRY_DELAY)

            with open(filepath, "rb") as img_file:
                response = requests.post(
                    API_URL,
                    files={"image": (filename, img_file, "image/jpeg")}
                )

            result = response.json()

            if result["success"]:
                d = result["data"]
                city = ""
                row = [
                    filename,
                    d.get("customer_name", ""),
                    d.get("consumer_number", ""),
                    d.get("discom", ""),
                    city,
                    d.get("monthly_units", ""),
                    d.get("bill_amount", ""),
                    d.get("per_unit_rate", ""),
                    d.get("billing_period", ""),
                    d.get("recommended_kw", ""),
                    d.get("monthly_savings_rs", ""),
                    d.get("system_cost_rs", ""),
                    d.get("payback_years", ""),
                    d.get("savings_25_years_rs", "")
                ]
                ws.append(row)
                wb.save(OUTPUT_FILE)  # Save after every bill
                success += 1
                print(f"✅ {filename} — ₹{d.get('bill_amount')} | {d.get('monthly_units')} units | {d.get('billing_period')}")
                time.sleep(5)  # small delay between each bill to avoid hitting rate limits
                processed = True
            else:
                error = result.get('error', '')
                if '503' in str(error) or '429' in str(error):
                    print(f"   ⚠️ {filename} — Gemini busy/quota, waiting {RETRY_DELAY}s...")
                    attempt += 1
                else:
                    print(f"❌ {filename} — API Error: {error}")
                    processed = True

        except Exception as e:
            print(f"❌ {filename} — Exception: {str(e)}")
            attempt += 1

    if not processed:
        failed += 1
        print(f"❌ {filename} — Failed after {MAX_RETRIES} attempts")

print(f"\n✅ Done! {success} new bills saved, {failed} failed.")
print(f"📊 Dataset saved to: {OUTPUT_FILE}")